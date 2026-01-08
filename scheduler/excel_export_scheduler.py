"""Excel export scheduler that runs in background thread."""

import logging
import threading
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional
import sqlite3

logger = logging.getLogger(__name__)


class ExcelExportScheduler:
    """Background scheduler for automatic Excel exports."""
    
    def __init__(self, db_path: str = "data/people_counter.db", exports_dir: str = "exports"):
        """
        Initialize Excel export scheduler.
        
        Args:
            db_path: Path to SQLite database file
            exports_dir: Base directory for export files
        """
        self.db_path = db_path
        self.exports_dir = Path(exports_dir)
        self.exports_dir.mkdir(exist_ok=True)
        
        # Create subdirectories
        self.daily_dir = self.exports_dir / "daily"
        self.summary_dir = self.exports_dir / "summary"
        self.daily_dir.mkdir(exist_ok=True)
        self.summary_dir.mkdir(exist_ok=True)
        
        self._running = False
        self._thread: Optional[threading.Thread] = None
        self._last_export_time: Optional[datetime] = None
        self._export_interval_seconds = 30 * 60  # 30 minutes
        
        logger.info(f"ExcelExportScheduler initialized: db={db_path}, exports_dir={exports_dir}")
    
    def start(self):
        """Start the background scheduler thread."""
        if self._running:
            logger.warning("ExcelExportScheduler already running")
            return
        
        self._running = True
        self._thread = threading.Thread(target=self._scheduler_loop, daemon=True)
        self._thread.start()
        logger.info("ExcelExportScheduler started")
        
        # Export immediately on startup (for testing)
        try:
            today = date.today().strftime('%Y-%m-%d')
            output_file = self.daily_dir / f"people_counter_{today}.xlsx"
            logger.info("Running immediate export on startup...")
            self._export_daily_excel(today, output_file)
            
            # Also export rolling summary on startup
            self._export_rolling_summary()
            
            self._last_export_time = datetime.now()
            logger.info("Immediate export completed")
        except Exception as e:
            logger.error(f"Error during immediate export: {e}", exc_info=True)
    
    def stop(self):
        """Stop the background scheduler thread."""
        if not self._running:
            return
        
        logger.info("Stopping ExcelExportScheduler...")
        self._running = False
        
        if self._thread:
            self._thread.join(timeout=5.0)
            if self._thread.is_alive():
                logger.warning("ExcelExportScheduler thread did not stop within timeout")
            else:
                logger.info("ExcelExportScheduler stopped")
    
    def _scheduler_loop(self):
        """Main scheduler loop running in background thread."""
        logger.info("ExcelExportScheduler thread started")
        
        while self._running:
            try:
                now = datetime.now()
                current_time = now.time()
                
                # Check if it's 00:00 (start of new day - finalize yesterday and run aggregated export)
                if current_time.hour == 0 and current_time.minute == 0:
                    # Finalize yesterday's file
                    yesterday = (date.today() - timedelta(days=1)).strftime('%Y-%m-%d')
                    yesterday_file = self.daily_dir / f"people_counter_{yesterday}.xlsx"
                    if not yesterday_file.exists():
                        # Export yesterday's final file
                        self._export_daily_excel(yesterday, yesterday_file)
                    
                    # Run rolling summary export (replaces old aggregated export)
                    self._export_rolling_summary()
                    
                    # Run cleanup
                    self._cleanup_old_files()
                    
                    # Wait until next minute to avoid multiple exports
                    time.sleep(60)
                    continue
                
                # Check if 1 minute has passed since last export (for testing)
                should_export = False
                if self._last_export_time is None:
                    should_export = True
                else:
                    elapsed = (now - self._last_export_time).total_seconds()
                    if elapsed >= self._export_interval_seconds:
                        should_export = True
                
                if should_export:
                    # Export to today's daily file
                    today = date.today().strftime('%Y-%m-%d')
                    output_file = self.daily_dir / f"people_counter_{today}.xlsx"
                    self._export_daily_excel(today, output_file)
                    
                    # Export rolling summary (last 5 days)
                    self._export_rolling_summary()
                    
                    self._last_export_time = now
                
                # Sleep for 1 minute before next check
                time.sleep(60)
                
            except Exception as e:
                logger.error(f"Error in ExcelExportScheduler loop: {e}", exc_info=True)
                time.sleep(60)  # Wait before retrying
        
        logger.info("ExcelExportScheduler thread stopped")
    
    def _export_daily_excel(self, target_date: str, output_file: Path) -> bool:
        """
        Export daily Excel report using new database-driven exporter.
        
        Args:
            target_date: Target date in YYYY-MM-DD format
            output_file: Output Excel file path (unused, kept for compatibility)
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Get morning times from config
            morning_start = "11:05"
            morning_end = "11:14"
            
            try:
                from app.config import load_config
                config = load_config()
                morning_start = config.production.morning_start
                morning_end = config.production.morning_end
                logger.debug(f"Using config morning times: {morning_start}-{morning_end}")
            except Exception as e:
                logger.warning(f"Could not load config, using defaults: {morning_start}-{morning_end} ({e})")
            
            # Use new database-driven exporter
            from export.excel_exporter import export_daily_excel
            
            result = export_daily_excel(
                target_date=target_date,
                db_path=self.db_path,
                output_dir=str(self.daily_dir),
                morning_start=morning_start,
                morning_end=morning_end
            )
            
            return result
            
        except Exception as e:
            logger.error(f"EXCEL_EXPORT_ERROR: {e}", exc_info=True)
            return False
        
        # Import here to avoid import errors if pandas/openpyxl not available
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.error("Required libraries not found: pandas, openpyxl")
            return False
        
        # Check if database exists
        if not Path(self.db_path).exists():
            logger.error(f"Database file not found: {self.db_path}")
            return False
        
        # Create temp file first
        temp_file = output_file.with_suffix('.tmp.xlsx')
        
        try:
            # Connect to database
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Create alert_logs table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS alert_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    alert_time TEXT NOT NULL,
                    expected_total INTEGER NOT NULL,
                    current_total INTEGER NOT NULL,
                    missing INTEGER NOT NULL,
                    created_at TEXT NOT NULL DEFAULT (datetime('now'))
                )
            """)
            conn.commit()
            
            # Get data
            summary = self._get_daily_summary(cursor, target_date)
            alerts = self._get_alerts_for_date(cursor, target_date)
            events = self._get_events_for_date(cursor, target_date)
            
            # Create Excel writer
            with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                # Sheet 1: SUMMARY
                summary_data = {
                    'Field': ['Date', 'Total Morning', 'Missing Periods', 'Last Updated'],
                    'Value': [
                        target_date,
                        summary['total_morning'] if summary else 0,
                        self._format_missing_periods(alerts),
                        summary['updated_at'] if summary and summary.get('updated_at') else 'N/A'
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='SUMMARY', index=False)
                
                # Sheet 2: ALERTS
                if alerts:
                    alerts_data = {
                        'Time Window': [self._format_time_for_display(a['alert_time']) for a in alerts],
                        'Expected': [a['expected_total'] for a in alerts],
                        'Current': [a['current_total'] for a in alerts],
                        'Missing': [a['missing'] for a in alerts]
                    }
                    df_alerts = pd.DataFrame(alerts_data)
                else:
                    df_alerts = pd.DataFrame(columns=['Time Window', 'Expected', 'Current', 'Missing'])
                df_alerts.to_excel(writer, sheet_name='ALERTS', index=False)
                
                # Sheet 3: EVENTS
                if events:
                    events_data = {
                        'Time': [self._format_time_for_display(e['event_time']) for e in events],
                        'Direction': [e['direction'].upper() for e in events],
                        'Camera': [e['camera_id'] for e in events]
                    }
                    df_events = pd.DataFrame(events_data)
                else:
                    df_events = pd.DataFrame(columns=['Time', 'Direction', 'Camera'])
                df_events.to_excel(writer, sheet_name='EVENTS', index=False)
            
            # Apply formatting
            wb = load_workbook(temp_file)
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for sheet_name in ['SUMMARY', 'ALERTS', 'EVENTS']:
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                
                # Format header row
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Freeze header row
                ws.freeze_panes = 'A2'
                
                # Enable filter
                ws.auto_filter.ref = ws.dimensions
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to temp file
            wb.save(temp_file)
            conn.close()
            
            # Rename temp file to final file (atomic operation)
            try:
                if output_file.exists():
                    # Try to remove existing file (might fail if open in Excel)
                    try:
                        output_file.unlink()
                    except PermissionError:
                        logger.warning(f"Cannot overwrite {output_file.name} - file may be open in Excel. Skipping export.")
                        temp_file.unlink()  # Clean up temp file
                        return False
                
                temp_file.rename(output_file)
                logger.info(f"Excel export completed: {output_file.name} ({len(events)} events, {len(alerts)} alerts)")
                return True
                
            except PermissionError:
                logger.warning(f"Cannot write {output_file.name} - file may be open in Excel. Skipping export.")
                temp_file.unlink()  # Clean up temp file
                return False
            except Exception as e:
                logger.error(f"Error renaming temp file: {e}")
                temp_file.unlink()  # Clean up temp file
                return False
                
        except PermissionError:
            logger.warning(f"Cannot write Excel file - file may be open. Skipping export.")
            if temp_file.exists():
                temp_file.unlink()
            return False
        except sqlite3.Error as e:
            logger.error(f"Database error during export: {e}")
            if temp_file.exists():
                temp_file.unlink()
            return False
        except Exception as e:
            logger.error(f"Error during Excel export: {e}", exc_info=True)
            if temp_file.exists():
                temp_file.unlink()
            return False
    
    def _get_daily_summary(self, cursor: sqlite3.Cursor, target_date: str) -> Optional[dict]:
        """Get daily summary for a specific date."""
        try:
            cursor.execute("""
                SELECT date, total_morning, updated_at
                FROM daily_summary
                WHERE date = ?
            """, (target_date,))
            row = cursor.fetchone()
            if row:
                return {
                    'date': row[0],
                    'total_morning': row[1],
                    'updated_at': row[2] if len(row) > 2 else None
                }
        except sqlite3.OperationalError:
            # Try last_updated column name
            try:
                cursor.execute("""
                    SELECT date, total_morning, last_updated
                    FROM daily_summary
                    WHERE date = ?
                """, (target_date,))
                row = cursor.fetchone()
                if row:
                    return {
                        'date': row[0],
                        'total_morning': row[1],
                        'updated_at': row[2] if len(row) > 2 else None
                    }
            except:
                pass
        return None
    
    def _get_alerts_for_date(self, cursor: sqlite3.Cursor, target_date: str) -> list:
        """Get alerts for a specific date."""
        try:
            cursor.execute("""
                SELECT alert_time, expected_total, current_total, missing
                FROM alert_logs
                WHERE date(alert_time) = ?
                ORDER BY alert_time ASC
            """, (target_date,))
            
            alerts = []
            for row in cursor.fetchall():
                alerts.append({
                    'alert_time': row[0],
                    'expected_total': row[1],
                    'current_total': row[2],
                    'missing': row[3]
                })
            return alerts
        except sqlite3.OperationalError:
            # Table might not exist yet
            return []
    
    def _get_events_for_date(self, cursor: sqlite3.Cursor, target_date: str) -> list:
        """Get events for a specific date."""
        cursor.execute("""
            SELECT event_time, direction, camera_id
            FROM people_events
            WHERE date(event_time) = ?
            ORDER BY event_time ASC
        """, (target_date,))
        
        events = []
        for row in cursor.fetchall():
            events.append({
                'event_time': row[0],
                'direction': row[1],
                'camera_id': row[2]
            })
        return events
    
    def _format_missing_periods(self, alerts: list) -> str:
        """Format alerts into missing periods string."""
        if not alerts:
            return "None"
        
        periods = []
        for alert in alerts:
            try:
                dt = datetime.fromisoformat(alert['alert_time'].replace('Z', '+00:00'))
                time_str = dt.strftime('%H:%M')
            except:
                time_str = alert['alert_time'][:5] if len(alert['alert_time']) >= 5 else alert['alert_time']
            
            missing = alert['missing']
            periods.append(f"{time_str} (-{missing})")
        
        return ", ".join(periods)
    
    def _format_time_for_display(self, iso_time: str) -> str:
        """Format ISO timestamp to readable time string."""
        try:
            dt = datetime.fromisoformat(iso_time.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return iso_time
    
    def _export_rolling_summary(self):
        """Export rolling summary Excel file (last 7 days per requirements)."""
        try:
            from export.rolling_summary_exporter import export_rolling_summary
            
            result = export_rolling_summary(
                daily_dir=str(self.daily_dir),
                summary_dir=str(self.summary_dir),
                max_days=7  # Changed from 5 to 7 per requirements
            )
            
            if result:
                logger.info("Rolling summary export completed successfully")
            else:
                logger.warning("Rolling summary export failed or skipped")
                
        except Exception as e:
            logger.error(f"Error during rolling summary export: {e}", exc_info=True)
    
    def _cleanup_old_files(self):
        """Delete daily Excel files older than 5 days."""
        try:
            cutoff_date = date.today() - timedelta(days=5)
            deleted_count = 0
            deleted_files = []
            
            for file_path in self.daily_dir.glob("people_counter_*.xlsx"):
                file_date = self._parse_date_from_filename(file_path.name)
                if file_date and file_date < cutoff_date:
                    try:
                        file_path.unlink()
                        deleted_count += 1
                        deleted_files.append(file_path.name)
                        logger.info(f"Deleted old file: {file_path.name}")
                    except Exception as e:
                        logger.warning(f"Could not delete {file_path.name}: {e}")
            
            if deleted_count > 0:
                logger.info(f"Cleanup completed: Deleted {deleted_count} old file(s): {', '.join(deleted_files)}")
            else:
                logger.debug("Cleanup: No old files to delete")
        except Exception as e:
            logger.error(f"Error during cleanup: {e}", exc_info=True)
    
    def _parse_date_from_filename(self, filename: str) -> Optional[date]:
        """Parse date from filename like 'people_counter_2026-01-05.xlsx'."""
        try:
            parts = filename.replace('.xlsx', '').split('_')
            if len(parts) >= 3:
                date_str = '_'.join(parts[-3:])
                return datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            pass
        return None

