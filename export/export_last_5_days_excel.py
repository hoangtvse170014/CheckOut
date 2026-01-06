"""Export aggregated Excel report for last 5 days."""

import sys
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Error: Required libraries not found. Please install:")
    print("  pip install pandas openpyxl")
    sys.exit(1)


def parse_date_from_filename(filename: str) -> Optional[date]:
    """Parse date from filename like 'people_counter_2026-01-05.xlsx'."""
    try:
        # Extract date part (YYYY-MM-DD)
        parts = filename.replace('.xlsx', '').split('_')
        if len(parts) >= 3:
            date_str = '_'.join(parts[-3:])  # Get last 3 parts (YYYY-MM-DD)
            return datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        pass
    return None


def get_latest_5_daily_files(daily_dir: Path) -> List[Path]:
    """Get the latest 5 daily Excel files sorted by date."""
    if not daily_dir.exists():
        return []
    
    files_with_dates = []
    for file_path in daily_dir.glob("people_counter_*.xlsx"):
        file_date = parse_date_from_filename(file_path.name)
        if file_date:
            files_with_dates.append((file_date, file_path))
    
    # Sort by date descending and take top 5
    files_with_dates.sort(key=lambda x: x[0], reverse=True)
    return [f[1] for f in files_with_dates[:5]]


def read_summary_from_excel(file_path: Path) -> Optional[Dict]:
    """Read SUMMARY sheet from Excel file."""
    try:
        df = pd.read_excel(file_path, sheet_name='SUMMARY')
        # Excel format: Field | Value columns
        summary = {}
        for _, row in df.iterrows():
            field = str(row.get('Field', ''))
            value = row.get('Value', '')
            if field == 'Date':
                summary['date'] = str(value) if pd.notna(value) else ''
            elif field == 'Total Morning':
                summary['total_morning'] = int(value) if pd.notna(value) else 0
            elif field == 'Missing Periods':
                summary['missing_periods'] = str(value) if pd.notna(value) else 'None'
        return summary if summary else None
    except Exception as e:
        print(f"Warning: Could not read SUMMARY from {file_path.name}: {e}")
        return None


def read_alerts_from_excel(file_path: Path) -> List[Dict]:
    """Read ALERTS sheet from Excel file."""
    try:
        df = pd.read_excel(file_path, sheet_name='ALERTS')
        alerts = []
        file_date = parse_date_from_filename(file_path.name)
        
        for _, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get('Time Window', '')) and pd.isna(row.get('Missing', '')):
                continue
            
            time_window = str(row.get('Time Window', '')) if pd.notna(row.get('Time Window', '')) else ''
            missing = int(row.get('Missing', 0)) if pd.notna(row.get('Missing', '')) else 0
            
            alerts.append({
                'date': file_date,
                'time_window': time_window,
                'missing': missing
            })
        return alerts
    except Exception as e:
        print(f"Warning: Could not read ALERTS from {file_path.name}: {e}")
        return []


def export_last_5_days_excel(daily_dir: str = "exports/daily", output_file: str = None) -> bool:
    """
    Export aggregated Excel report for last 5 days.
    
    Args:
        daily_dir: Directory containing daily Excel files
        output_file: Output Excel file path (default: exports/summary/people_counter_last_5_days.xlsx)
    
    Returns:
        True if successful, False otherwise
    """
    daily_path = Path(daily_dir)
    if output_file is None:
        summary_dir = daily_path.parent / "summary"
        summary_dir.mkdir(exist_ok=True)
        output_file = summary_dir / "people_counter_last_5_days.xlsx"
    else:
        output_file = Path(output_file)
        output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Get latest 5 daily files
    daily_files = get_latest_5_daily_files(daily_path)
    
    if not daily_files:
        print(f"No daily Excel files found in {daily_dir}")
        return False
    
    print(f"Found {len(daily_files)} daily files for aggregation")
    
    # Create temp file
    temp_file = output_file.with_suffix('.tmp.xlsx')
    
    try:
        # Collect data from all files
        overview_data = []
        alert_history_data = []
        
        for file_path in daily_files:
            # Read SUMMARY
            summary = read_summary_from_excel(file_path)
            if summary:
                overview_data.append({
                    'Date': summary.get('date', ''),
                    'Total Morning': summary.get('total_morning', 0),
                    'Missing Periods': summary.get('missing_periods', 'None')
                })
            
            # Read ALERTS
            alerts = read_alerts_from_excel(file_path)
            for alert in alerts:
                date_str = ''
                if alert['date']:
                    if isinstance(alert['date'], date):
                        date_str = alert['date'].strftime('%Y-%m-%d')
                    else:
                        date_str = str(alert['date'])
                alert_history_data.append({
                    'Date': date_str,
                    'Time Window': alert['time_window'],
                    'Missing': alert['missing']
                })
        
        # Create Excel
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            # Sheet 1: OVERVIEW
            if overview_data:
                df_overview = pd.DataFrame(overview_data)
            else:
                df_overview = pd.DataFrame(columns=['Date', 'Total Morning', 'Missing Periods'])
            df_overview.to_excel(writer, sheet_name='OVERVIEW', index=False)
            
            # Sheet 2: ALERT_HISTORY
            if alert_history_data:
                df_alerts = pd.DataFrame(alert_history_data)
            else:
                df_alerts = pd.DataFrame(columns=['Date', 'Time Window', 'Missing'])
            df_alerts.to_excel(writer, sheet_name='ALERT_HISTORY', index=False)
        
        # Apply formatting
        wb = load_workbook(temp_file)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for sheet_name in ['OVERVIEW', 'ALERT_HISTORY']:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Enable filter
            ws.auto_filter.ref = ws.dimensions
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temp file
        wb.save(temp_file)
        
        # Rename temp file to final file
        try:
            if output_file.exists():
                output_file.unlink()
            temp_file.rename(output_file)
            print(f"\nAggregated export completed successfully!")
            print(f"Output file: {output_file}")
            print(f"Days aggregated: {len(daily_files)}")
            return True
        except PermissionError:
            print(f"Warning: Cannot overwrite {output_file.name} - file may be open in Excel")
            temp_file.unlink()
            return False
        except Exception as e:
            print(f"Error renaming temp file: {e}")
            temp_file.unlink()
            return False
            
    except Exception as e:
        print(f"Error during aggregated export: {e}")
        import traceback
        traceback.print_exc()
        if temp_file.exists():
            temp_file.unlink()
        return False


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Export aggregated Excel report for last 5 days")
    parser.add_argument(
        "--daily-dir",
        default="exports/daily",
        help="Directory containing daily Excel files (default: exports/daily)"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel file path (default: exports/summary/people_counter_last_5_days.xlsx)"
    )
    
    args = parser.parse_args()
    export_last_5_days_excel(args.daily_dir, args.output)

