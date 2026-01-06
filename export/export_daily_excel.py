"""Export daily Excel report from SQLite database."""

import sqlite3
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional, List, Tuple

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("Error: Required libraries not found. Please install:")
    print("  pip install pandas openpyxl")
    sys.exit(1)


def create_alert_logs_table_if_not_exists(cursor: sqlite3.Cursor):
    """Create alert_logs table if it doesn't exist."""
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS alert_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alert_time TEXT NOT NULL,
            expected_total INTEGER NOT NULL,
            current_total INTEGER NOT NULL,
            missing INTEGER NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)


def get_daily_summary(cursor: sqlite3.Cursor, target_date: str) -> Optional[dict]:
    """Get daily summary for a specific date."""
    cursor.execute("""
        SELECT date, total_morning, updated_at
        FROM daily_summary
        WHERE date = ?
    """, (target_date,))
    row = cursor.fetchone()
    
    if row:
        return {
            'date': row[0],
            'total_morning': row[1],
            'updated_at': row[2] if len(row) > 2 else None
        }
    return None


def get_daily_state(cursor: sqlite3.Cursor, target_date: str) -> Optional[dict]:
    """Get daily state for a specific date."""
    cursor.execute("""
        SELECT date, total_morning, realtime_in, realtime_out, updated_at
        FROM daily_state
        WHERE date = ?
    """, (target_date,))
    row = cursor.fetchone()
    
    if row:
        return {
            'date': row[0],
            'total_morning': row[1] if row[1] is not None else 0,
            'realtime_in': row[2] if row[2] is not None else 0,
            'realtime_out': row[3] if row[3] is not None else 0,
            'updated_at': row[4] if len(row) > 4 else None
        }
    return None


def get_alerts_for_date(cursor: sqlite3.Cursor, target_date: str) -> List[dict]:
    """Get alerts for a specific date."""
    # Filter alerts by date (assuming alert_time is in ISO format)
    cursor.execute("""
        SELECT alert_time, expected_total, current_total, missing
        FROM alert_logs
        WHERE date(alert_time) = ?
        ORDER BY alert_time ASC
    """, (target_date,))
    
    alerts = []
    for row in cursor.fetchall():
        alerts.append({
            'alert_time': row[0],
            'expected_total': row[1],
            'current_total': row[2],
            'missing': row[3]
        })
    return alerts


def calculate_missing_periods_with_duration(alerts: List[dict]) -> List[dict]:
    """
    Calculate missing periods with start time, end time, and duration.
    Groups consecutive alerts into periods.
    """
    if not alerts:
        return []
    
    periods = []
    current_period = None
    
    for alert in alerts:
        try:
            alert_dt = datetime.fromisoformat(alert['alert_time'].replace('Z', '+00:00'))
        except:
            continue
        
        if current_period is None:
            # Start new period
            current_period = {
                'start_time': alert_dt,
                'end_time': alert_dt,
                'missing': alert['missing'],
                'expected': alert['expected_total'],
                'current': alert['current_total']
            }
        else:
            # Check if this alert is within 30 minutes of the previous one (same period)
            time_diff = (alert_dt - current_period['end_time']).total_seconds() / 60  # minutes
            
            if time_diff <= 30 and alert['missing'] == current_period['missing']:
                # Extend current period
                current_period['end_time'] = alert_dt
            else:
                # Save current period and start new one
                duration = (current_period['end_time'] - current_period['start_time']).total_seconds() / 60
                periods.append({
                    'start_time': current_period['start_time'],
                    'end_time': current_period['end_time'],
                    'duration_minutes': duration,
                    'missing': current_period['missing'],
                    'expected': current_period['expected'],
                    'current': current_period['current']
                })
                current_period = {
                    'start_time': alert_dt,
                    'end_time': alert_dt,
                    'missing': alert['missing'],
                    'expected': alert['expected_total'],
                    'current': alert['current_total']
                }
    
    # Save last period
    if current_period:
        duration = (current_period['end_time'] - current_period['start_time']).total_seconds() / 60
        periods.append({
            'start_time': current_period['start_time'],
            'end_time': current_period['end_time'],
            'duration_minutes': duration,
            'missing': current_period['missing'],
            'expected': current_period['expected'],
            'current': current_period['current']
        })
    
    return periods


def format_missing_periods(alerts: List[dict]) -> str:
    """Format alerts into missing periods string."""
    if not alerts:
        return "None"
    
    periods = []
    for alert in alerts:
        # Extract time from ISO timestamp
        try:
            dt = datetime.fromisoformat(alert['alert_time'].replace('Z', '+00:00'))
            time_str = dt.strftime('%H:%M')
        except:
            time_str = alert['alert_time'][:5] if len(alert['alert_time']) >= 5 else alert['alert_time']
        
        missing = alert['missing']
        periods.append(f"{time_str} (-{missing})")
    
    return ", ".join(periods)


def calculate_total_morning_from_events(cursor: sqlite3.Cursor, target_date: str, morning_start: str, morning_end: str) -> int:
    """
    Calculate total_morning from events in morning phase.
    
    Args:
        cursor: Database cursor
        target_date: Target date (YYYY-MM-DD)
        morning_start: Morning phase start (HH:MM)
        morning_end: Morning phase end (HH:MM)
    
    Returns:
        Total morning count (IN - OUT during morning phase)
    """
    try:
        start_hour, start_min = map(int, morning_start.split(':'))
        end_hour, end_min = map(int, morning_end.split(':'))
        
        cursor.execute("""
            SELECT direction, COUNT(*) as count
            FROM people_events
            WHERE date(event_time) = ?
              AND CAST(strftime('%H', event_time) AS INTEGER) * 60 + CAST(strftime('%M', event_time) AS INTEGER) >= ?
              AND CAST(strftime('%H', event_time) AS INTEGER) * 60 + CAST(strftime('%M', event_time) AS INTEGER) < ?
            GROUP BY direction
        """, (target_date, start_hour * 60 + start_min, end_hour * 60 + end_min))
        
        results = cursor.fetchall()
        in_count = 0
        out_count = 0
        
        for direction, count in results:
            if direction == 'IN':
                in_count = count
            elif direction == 'OUT':
                out_count = count
        
        return in_count - out_count
    except Exception as e:
        print(f"Error calculating total_morning from events: {e}")
        return 0


def get_events_for_date(cursor: sqlite3.Cursor, target_date: str) -> List[dict]:
    """Get events for a specific date from people_events table."""
    # Try to query with date() function first
    try:
        cursor.execute("""
            SELECT event_time, direction, camera_id
            FROM people_events
            WHERE date(event_time) = ?
            ORDER BY event_time ASC
        """, (target_date,))
    except sqlite3.OperationalError:
        # Fallback: if date() doesn't work, use LIKE pattern
        cursor.execute("""
            SELECT event_time, direction, camera_id
            FROM people_events
            WHERE event_time LIKE ?
            ORDER BY event_time ASC
        """, (f"{target_date}%",))
    
    events = []
    for row in cursor.fetchall():
        events.append({
            'event_time': row[0],
            'direction': row[1],
            'camera_id': row[2] if len(row) > 2 else ''
        })
    return events


def get_statistics_from_events(events: List[dict]) -> dict:
    """Calculate statistics from events."""
    stats = {
        'total_in': 0,
        'total_out': 0,
        'first_event_time': None,
        'last_event_time': None
    }
    
    for event in events:
        if event['direction'].lower() == 'in':
            stats['total_in'] += 1
        elif event['direction'].lower() == 'out':
            stats['total_out'] += 1
        
        try:
            event_dt = datetime.fromisoformat(event['event_time'].replace('Z', '+00:00'))
            if stats['first_event_time'] is None or event_dt < stats['first_event_time']:
                stats['first_event_time'] = event_dt
            if stats['last_event_time'] is None or event_dt > stats['last_event_time']:
                stats['last_event_time'] = event_dt
        except:
            pass
    
    return stats


def format_time_for_display(iso_time: str) -> str:
    """Format ISO timestamp to readable time string."""
    try:
        dt = datetime.fromisoformat(iso_time.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except:
        return iso_time


def export_daily_excel(target_date: Optional[str] = None, db_path: str = "data/people_counter.db") -> bool:
    """
    Export daily Excel report from SQLite database.
    
    IMPORTANT: db_path must match the database path used by the app.
    Default is "data/people_counter.db" to match app config.
    """
    """
    Export daily Excel report from SQLite database.
    
    Args:
        target_date: Target date in YYYY-MM-DD format (default: today)
        db_path: Path to SQLite database file
    
    Returns:
        True if successful, False otherwise
    """
    # Set target date
    if target_date is None:
        target_date = date.today().strftime('%Y-%m-%d')
    
    # Validate date format
    try:
        datetime.strptime(target_date, '%Y-%m-%d')
    except ValueError:
        print(f"Error: Invalid date format. Use YYYY-MM-DD")
        return False
    
    # Check if database exists
    if not Path(db_path).exists():
        print(f"Error: Database file not found: {db_path}")
        return False
    
    # Create exports directory (use daily subdirectory)
    exports_dir = Path("exports") / "daily"
    exports_dir.mkdir(parents=True, exist_ok=True)
    
    # Output file path (use temp file first, then rename to avoid permission errors)
    output_file = exports_dir / f"people_counter_{target_date}.xlsx"
    temp_file = exports_dir / f"people_counter_{target_date}.tmp.xlsx"
    
    try:
        # Connect to database
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Create tables if they don't exist
        create_alert_logs_table_if_not_exists(cursor)
        
        # Create people_events table if it doesn't exist (new schema)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS people_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_time TEXT NOT NULL,
                direction TEXT NOT NULL,
                camera_id TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        
        # Create daily_summary table if it doesn't exist
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_summary (
                date TEXT PRIMARY KEY,
                total_morning INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        
        # Also check for old schema (events table) and migrate if needed
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='events'")
        has_old_events = cursor.fetchone() is not None
        
        conn.commit()
        
        # Get data - try both schemas
        summary = get_daily_summary(cursor, target_date)
        daily_state = get_daily_state(cursor, target_date)
        alerts = get_alerts_for_date(cursor, target_date)
        
        # Try new schema first (people_events), then old schema (events)
        events = get_events_for_date(cursor, target_date)
        
        # Calculate total_morning from morning phase events if not in daily_state
        # Default morning phase: 07:00-08:40 (can be overridden via config)
        morning_start = "07:00"
        morning_end = "08:40"
        
        # Try to get from config if available (read from env or use defaults)
        import os
        morning_start_env = os.getenv("PRODUCTION__MORNING_START", morning_start)
        morning_end_env = os.getenv("PRODUCTION__MORNING_END", morning_end)
        if morning_start_env and morning_end_env:
            morning_start = morning_start_env
            morning_end = morning_end_env
        if not events and has_old_events:
            # Try old schema
            cursor.execute("""
                SELECT timestamp as event_time, direction, camera_id
                FROM events
                WHERE date(timestamp) = ?
                ORDER BY timestamp ASC
            """, (target_date,))
            events = []
            for row in cursor.fetchall():
                events.append({
                    'event_time': row[0],
                    'direction': row[1],
                    'camera_id': row[2]
                })
        
        # Check if database is empty and show warning
        has_data = (summary and summary.get('total_morning', 0) > 0) or (daily_state and daily_state.get('total_morning', 0) > 0) or len(events) > 0 or len(alerts) > 0
        stats = get_statistics_from_events(events)
        missing_periods = calculate_missing_periods_with_duration(alerts)
        
        # Prepare summary data - use daily_state if available (more accurate)
        # Debug: Print what we found
        print(f"\nDebug - daily_state: {daily_state}")
        print(f"Debug - summary: {summary}")
        print(f"Debug - events count: {len(events)}")
        
        if daily_state:
            total_morning = daily_state.get('total_morning', 0) or 0
            # If total_morning is 0 or None, calculate from morning phase events
            if total_morning == 0 or total_morning is None:
                total_morning = calculate_total_morning_from_events(cursor, target_date, morning_start, morning_end)
                if total_morning > 0:
                    print(f"Debug - Calculated total_morning from events: {total_morning}")
            realtime_in = daily_state.get('realtime_in', 0) or 0
            realtime_out = daily_state.get('realtime_out', 0) or 0
            # If no events, use realtime counts from state
            if not events:
                total_in_all_day = realtime_in
                total_out_all_day = realtime_out
                status_msg = "Data from State (No Events Recorded)"
            else:
                total_in_all_day = stats['total_in']
                total_out_all_day = stats['total_out']
                status_msg = "OK"
            last_updated = daily_state.get('updated_at', 'N/A')
            print(f"Debug - Using daily_state: total_morning={total_morning}, realtime_in={realtime_in}, realtime_out={realtime_out}")
        elif summary:
            total_morning = summary.get('total_morning', 0) or 0
            # If total_morning is 0, calculate from morning phase events
            if total_morning == 0:
                total_morning = calculate_total_morning_from_events(cursor, target_date, morning_start, morning_end)
                if total_morning > 0:
                    print(f"Debug - Calculated total_morning from events: {total_morning}")
            total_in_all_day = stats['total_in']
            total_out_all_day = stats['total_out']
            last_updated = summary.get('updated_at', 'N/A')
            status_msg = "OK" if has_data else "Database Empty"
            print(f"Debug - Using summary: total_morning={total_morning}")
        else:
            # Calculate from morning phase events as last resort
            total_morning = calculate_total_morning_from_events(cursor, target_date, morning_start, morning_end)
            total_in_all_day = stats['total_in']
            total_out_all_day = stats['total_out']
            last_updated = 'N/A'
            status_msg = "Database Empty / No Data for this Date"
            print(f"Debug - No state found, using defaults")
        
        # Create Excel writer (use temp file first)
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            # ========== Sheet 1: SUMMARY ==========
            first_event_str = stats['first_event_time'].strftime('%H:%M:%S') if stats['first_event_time'] else 'N/A'
            last_event_str = stats['last_event_time'].strftime('%H:%M:%S') if stats['last_event_time'] else 'N/A'
            
            summary_data = {
                'Field': [
                    'Date', 
                    'Total Morning', 
                    'Total IN (All Day)', 
                    'Total OUT (All Day)',
                    'First Event Time',
                    'Last Event Time',
                    'Missing Periods Count',
                    'Last Updated',
                    'Status'
                ],
                'Value': [
                    target_date,
                    total_morning,
                    total_in_all_day,
                    total_out_all_day,
                    first_event_str,
                    last_event_str,
                    len(missing_periods),
                    last_updated,
                    status_msg
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='SUMMARY', index=False)
            
            # ========== Sheet 2: MISSING PERIODS (with duration) ==========
            if missing_periods:
                periods_data = {
                    'Start Time': [p['start_time'].strftime('%Y-%m-%d %H:%M:%S') for p in missing_periods],
                    'End Time': [p['end_time'].strftime('%Y-%m-%d %H:%M:%S') for p in missing_periods],
                    'Duration (minutes)': [round(p['duration_minutes'], 1) for p in missing_periods],
                    'Duration (formatted)': [
                        f"{int(p['duration_minutes']//60)}h {int(p['duration_minutes']%60)}m" 
                        if p['duration_minutes'] >= 60 
                        else f"{int(p['duration_minutes'])}m" 
                        for p in missing_periods
                    ],
                    'Missing People': [p['missing'] for p in missing_periods],
                    'Expected': [p['expected'] for p in missing_periods],
                    'Current': [p['current'] for p in missing_periods]
                }
                df_periods = pd.DataFrame(periods_data)
            else:
                df_periods = pd.DataFrame(columns=[
                    'Start Time', 'End Time', 'Duration (minutes)', 'Duration (formatted)', 
                    'Missing People', 'Expected', 'Current'
                ])
            df_periods.to_excel(writer, sheet_name='MISSING PERIODS', index=False)
            
            # ========== Sheet 3: ALERTS (detailed) ==========
            if alerts:
                alerts_data = {
                    'Time': [format_time_for_display(a['alert_time']) for a in alerts],
                    'Expected': [a['expected_total'] for a in alerts],
                    'Current': [a['current_total'] for a in alerts],
                    'Missing': [a['missing'] for a in alerts]
                }
                df_alerts = pd.DataFrame(alerts_data)
            else:
                df_alerts = pd.DataFrame(columns=['Time', 'Expected', 'Current', 'Missing'])
            df_alerts.to_excel(writer, sheet_name='ALERTS', index=False)
            
            # ========== Sheet 4: EVENTS (IN/OUT with time) ==========
            if events:
                events_data = {
                    'Time': [format_time_for_display(e['event_time']) for e in events],
                    'Direction': [e['direction'].upper() for e in events],
                    'Camera': [e['camera_id'] for e in events]
                }
                df_events = pd.DataFrame(events_data)
            else:
                df_events = pd.DataFrame(columns=['Time', 'Direction', 'Camera'])
            df_events.to_excel(writer, sheet_name='EVENTS', index=False)
        
        # Apply formatting using openpyxl (load from temp_file, not output_file)
        wb = load_workbook(temp_file)
        
        # Format all sheets
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for sheet_name in ['SUMMARY', 'MISSING PERIODS', 'ALERTS', 'EVENTS']:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Enable filter
            ws.auto_filter.ref = ws.dimensions
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook to temp file first
        wb.save(temp_file)
        conn.close()
        
        # Rename temp file to final file (atomic operation)
        try:
            if output_file.exists():
                # Try to remove existing file (might fail if open in Excel)
                try:
                    output_file.unlink()
                except PermissionError:
                    print(f"Warning: Cannot overwrite {output_file.name} - file may be open in Excel. Skipping export.")
                    temp_file.unlink()  # Clean up temp file
                    return False
            
            temp_file.rename(output_file)
            
            print(f"\nExport completed successfully!")
            print(f"Date: {target_date}")
            print(f"Total Morning: {total_morning}")
            print(f"Total IN: {total_in_all_day}, Total OUT: {total_out_all_day}")
            print(f"Events: {len(events)}")
            print(f"Missing Periods: {len(missing_periods)}")
            print(f"Alerts: {len(alerts)}")
            print(f"Output file: {output_file}")
            return True
            
        except PermissionError:
            print(f"Warning: Cannot write {output_file.name} - file may be open in Excel. Skipping export.")
            temp_file.unlink()  # Clean up temp file
            return False
        except Exception as e:
            print(f"Error renaming temp file: {e}")
            temp_file.unlink()  # Clean up temp file
            return False
        
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return False
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Export daily Excel report from SQLite")
    parser.add_argument(
        "--date",
        default=None,
        help="Target date in YYYY-MM-DD format (default: today)"
    )
    parser.add_argument(
        "--db",
        default="data/people_counter.db",
        help="Path to SQLite database file (default: data/people_counter.db)"
    )
    
    args = parser.parse_args()
    export_daily_excel(args.date, args.db)

