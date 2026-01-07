"""
Rolling summary exporter - creates people_counter_LAST_5_DAYS.xlsx from daily Excel files.
Reads only from existing daily Excel files (trusted source).
"""

import logging
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from export.retention_manager import get_valid_daily_files, cleanup_old_daily_files

logger = logging.getLogger(__name__)


def export_rolling_summary(
    daily_dir: str = "exports/daily",
    summary_dir: str = "exports/summary",
    max_days: int = 5
) -> bool:
    """
    Export rolling summary Excel file with last 5 days of data.
    
    Process:
    1. Cleanup old daily files (keep only max_days)
    2. Get valid daily Excel files
    3. Read data from each daily file
    4. Aggregate into summary sheets
    5. Write to summary Excel file (atomic write)
    
    Args:
        daily_dir: Directory containing daily Excel files
        summary_dir: Directory for summary Excel file
        max_days: Maximum number of days to include (default: 5)
    
    Returns:
        True if successful, False otherwise
    """
    try:
        # Step 1: Cleanup old files first
        deleted_count, deleted_files = cleanup_old_daily_files(daily_dir, max_days)
        if deleted_count > 0:
            logger.info(f"Cleaned up {deleted_count} old daily file(s)")
        
        # Step 2: Get valid daily files
        daily_files = get_valid_daily_files(daily_dir, max_days)
        
        if not daily_files:
            logger.warning("No daily Excel files found for summary")
            return False
        
        logger.info(f"Building summary from {len(daily_files)} daily file(s)")
        
        # Step 3: Read data from daily files
        daily_summaries = []
        all_alerts = []
        all_missing_periods = []
        
        for file_date, file_path in daily_files:
            try:
                data = _read_daily_file(file_path, file_date)
                if data:
                    daily_summaries.append(data['summary'])
                    all_alerts.extend(data['alerts'])
                    all_missing_periods.extend(data['missing_periods'])
            except Exception as e:
                logger.error(f"Error reading daily file {file_path.name}: {e}", exc_info=True)
                continue
        
        if not daily_summaries:
            logger.error("No valid data found in daily files")
            return False
        
        # Step 4: Setup output paths
        summary_path = Path(summary_dir)
        summary_path.mkdir(parents=True, exist_ok=True)
        
        output_file = summary_path / "people_counter_LAST_5_DAYS.xlsx"
        temp_file = summary_path / "people_counter_LAST_5_DAYS.tmp.xlsx"
        
        # Remove temp file if exists
        if temp_file.exists():
            try:
                temp_file.unlink()
            except Exception as e:
                logger.warning(f"Could not remove old temp file: {e}")
        
        logger.info(f"ROLLING_SUMMARY_START: Building summary from {len(daily_summaries)} day(s)")
        
        # Step 5: Create Excel file
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            # Sheet 1: DAILY_SUMMARY
            df_daily_summary = pd.DataFrame(daily_summaries)
            df_daily_summary.to_excel(writer, sheet_name='DAILY_SUMMARY', index=False)
            
            # Sheet 2: DAILY_ALERTS
            if all_alerts:
                df_alerts = pd.DataFrame(all_alerts)
            else:
                df_alerts = pd.DataFrame(columns=['Date', 'alert_time', 'total_morning', 'realtime', 'missing'])
            df_alerts.to_excel(writer, sheet_name='DAILY_ALERTS', index=False)
            
            # Sheet 3: DAILY_MISSING_PERIODS
            if all_missing_periods:
                df_missing = pd.DataFrame(all_missing_periods)
            else:
                df_missing = pd.DataFrame(columns=['Date', 'start_time', 'end_time', 'duration_minutes'])
            df_missing.to_excel(writer, sheet_name='DAILY_MISSING_PERIODS', index=False)
        
        # Step 6: Format Excel file
        wb = load_workbook(temp_file)
        _format_summary_excel(wb)
        wb.save(temp_file)
        
        # Step 7: Atomic replace
        try:
            if output_file.exists():
                output_file.unlink()
            
            temp_file.rename(output_file)
            
            logger.info(
                f"ROLLING_SUMMARY_SUCCESS: {output_file.name} "
                f"({len(daily_summaries)} days, {len(all_alerts)} alerts, "
                f"{len(all_missing_periods)} missing periods)"
            )
            return True
            
        except PermissionError:
            logger.warning(
                f"ROLLING_SUMMARY_SKIPPED: Cannot overwrite {output_file.name} - "
                f"file may be open. Temp file preserved: {temp_file.name}"
            )
            return False
        except Exception as e:
            logger.error(f"ROLLING_SUMMARY_ERROR: Failed to rename temp file: {e}")
            if temp_file.exists():
                temp_file.unlink()
            return False
            
    except Exception as e:
        logger.error(f"ROLLING_SUMMARY_ERROR: {e}", exc_info=True)
        if 'temp_file' in locals() and temp_file.exists():
            temp_file.unlink()
        return False


def _read_daily_file(file_path: Path, file_date) -> Optional[Dict]:
    """
    Read data from a daily Excel file.
    
    Args:
        file_path: Path to daily Excel file
        file_date: date object for the file
    
    Returns:
        Dict with 'summary', 'alerts', 'missing_periods' or None if error
    """
    try:
        date_str = file_date.strftime('%Y-%m-%d')
        
        # Read all sheets
        xls = pd.ExcelFile(file_path)
        
        # Read SUMMARY sheet
        if 'SUMMARY' not in xls.sheet_names:
            logger.warning(f"SUMMARY sheet not found in {file_path.name}")
            return None
        
        df_summary = pd.read_excel(xls, sheet_name='SUMMARY')
        summary_dict = dict(zip(df_summary['Field'], df_summary['Value']))
        
        # Extract summary data
        total_morning = int(summary_dict.get('Total Morning', 0) or 0)
        realtime = int(summary_dict.get('Current Realtime', 0) or 0)
        
        # Read ALERTS sheet
        alerts = []
        if 'ALERTS' in xls.sheet_names:
            df_alerts = pd.read_excel(xls, sheet_name='ALERTS')
            if not df_alerts.empty:
                for _, row in df_alerts.iterrows():
                    alerts.append({
                        'Date': date_str,
                        'alert_time': str(row.get('alert_time', '')),
                        'total_morning': int(row.get('total_morning', 0) or 0),
                        'realtime': int(row.get('realtime', 0) or 0),
                        'missing': int(row.get('missing', 0) or 0)
                    })
        
        # Read MISSING_PERIODS sheet
        missing_periods = []
        if 'MISSING_PERIODS' in xls.sheet_names:
            df_missing = pd.read_excel(xls, sheet_name='MISSING_PERIODS')
            if not df_missing.empty:
                total_missing_minutes = 0
                for _, row in df_missing.iterrows():
                    duration = int(row.get('duration_minutes', 0) or 0)
                    total_missing_minutes += duration
                    missing_periods.append({
                        'Date': date_str,
                        'start_time': str(row.get('start_time', '')),
                        'end_time': str(row.get('end_time', '')),
                        'duration_minutes': duration
                    })
        else:
            total_missing_minutes = 0
        
        # Read EVENTS sheet to calculate max/min realtime
        max_realtime = realtime
        min_realtime = realtime
        
        if 'EVENTS' in xls.sheet_names:
            df_events = pd.read_excel(xls, sheet_name='EVENTS')
            if not df_events.empty:
                # Calculate cumulative count from events
                current_count = 0
                max_realtime = 0
                min_realtime = 0
                
                for _, row in df_events.iterrows():
                    direction = str(row.get('direction', '')).upper()
                    if direction == 'IN':
                        current_count += 1
                    elif direction == 'OUT':
                        current_count -= 1
                    
                    max_realtime = max(max_realtime, current_count)
                    min_realtime = min(min_realtime, current_count)
        
        # Build summary row
        summary = {
            'Date': date_str,
            'Total Morning': total_morning,
            'Max Realtime': max_realtime,
            'Min Realtime': min_realtime,
            'Final Realtime': realtime,
            'Total Alerts': len(alerts),
            'Total Missing Periods': len(missing_periods),
            'Total Missing Minutes': total_missing_minutes
        }
        
        return {
            'summary': summary,
            'alerts': alerts,
            'missing_periods': missing_periods
        }
        
    except Exception as e:
        logger.error(f"Error reading daily file {file_path.name}: {e}", exc_info=True)
        return None


def _format_summary_excel(wb):
    """Apply formatting to summary Excel workbook."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for sheet_name in ['DAILY_SUMMARY', 'DAILY_ALERTS', 'DAILY_MISSING_PERIODS']:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Enable filter
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

