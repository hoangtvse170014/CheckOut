"""
Excel exporter - 100% database-driven, atomic writes.
SQLite is the single source of truth.
"""

import logging
import sqlite3
from pathlib import Path
from typing import Optional
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from export.db_queries import get_all_data_for_date

logger = logging.getLogger(__name__)


def export_daily_excel(
    target_date: str,
    db_path: str,
    output_dir: str = "exports/daily",
    morning_start: str = "06:00",
    morning_end: str = "08:30"
) -> bool:
    """
    Export daily Excel file - completely rebuilt from database.
    
    This function:
    1. Queries SQLite database directly (no memory state)
    2. Rebuilds Excel file completely from scratch
    3. Uses atomic write (temp file -> rename)
    4. Never depends on in-memory counters
    
    Args:
        target_date: Date in YYYY-MM-DD format
        db_path: Path to SQLite database
        output_dir: Output directory for Excel files
        morning_start: Morning phase start time (HH:MM)
        morning_end: Morning phase end time (HH:MM)
    
    Returns:
        True if successful, False otherwise
    """
    try:
        # Validate inputs
        if not Path(db_path).exists():
            logger.error(f"Database file not found: {db_path}")
            return False
        
        # Setup paths
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        output_file = output_path / f"people_counter_{target_date}.xlsx"
        temp_file = output_path / f"people_counter_{target_date}.tmp.xlsx"
        
        # Remove temp file if exists (cleanup from previous failed export)
        if temp_file.exists():
            try:
                temp_file.unlink()
            except Exception as e:
                logger.warning(f"Could not remove old temp file: {e}")
        
        logger.info(f"EXCEL_EXPORT_START: date={target_date}, db={db_path}")
        
        # Connect to database
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        try:
            # Get all data from database
            data = get_all_data_for_date(cursor, target_date, morning_start, morning_end)
            
            logger.info(
                f"Data retrieved: total_morning={data['total_morning']}, "
                f"realtime={data['realtime']}, missing={data['missing']}, "
                f"events={len(data['events'])}, alerts={len(data['alerts'])}, "
                f"missing_periods={len(data['missing_periods'])}"
            )
            
            # Create Excel file in temp location
            with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                # Sheet 1: SUMMARY
                summary_data = {
                    'Field': [
                        'Date',
                        'Total Morning',
                        'Current Realtime',
                        'Current Missing',
                        'Last Updated Time'
                    ],
                    'Value': [
                        target_date,
                        data['total_morning'],
                        data['realtime'],
                        data['missing'],
                        data['last_updated']
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='SUMMARY', index=False)
                
                # Sheet 2: MISSING_PERIODS
                if data['missing_periods']:
                    missing_periods_data = {
                        'start_time': [p['start_time'] for p in data['missing_periods']],
                        'end_time': [p.get('end_time', '') for p in data['missing_periods']],
                        'duration_minutes': [p.get('duration_minutes', 0) for p in data['missing_periods']],
                        'session': [p.get('session', '') for p in data['missing_periods']]
                    }
                    df_missing = pd.DataFrame(missing_periods_data)
                else:
                    df_missing = pd.DataFrame(columns=['start_time', 'end_time', 'duration_minutes', 'session'])
                df_missing.to_excel(writer, sheet_name='MISSING_PERIODS', index=False)
                
                # Sheet 3: ALERTS
                if data['alerts']:
                    alerts_data = {
                        'alert_time': [a['alert_time'] for a in data['alerts']],
                        'total_morning': [a['total_morning'] for a in data['alerts']],
                        'realtime': [a['realtime'] for a in data['alerts']],
                        'missing': [a['missing'] for a in data['alerts']]
                    }
                    df_alerts = pd.DataFrame(alerts_data)
                else:
                    df_alerts = pd.DataFrame(columns=['alert_time', 'total_morning', 'realtime', 'missing'])
                df_alerts.to_excel(writer, sheet_name='ALERTS', index=False)
                
                # Sheet 4: EVENTS
                if data['events']:
                    events_data = {
                        'event_time': [e['event_time'] for e in data['events']],
                        'direction': [e['direction'] for e in data['events']],
                        'camera_id': [e['camera_id'] for e in data['events']]
                    }
                    df_events = pd.DataFrame(events_data)
                else:
                    df_events = pd.DataFrame(columns=['event_time', 'direction', 'camera_id'])
                df_events.to_excel(writer, sheet_name='EVENTS', index=False)
            
            # Format Excel file
            wb = load_workbook(temp_file)
            _format_excel(wb)
            wb.save(temp_file)
            
            # Atomic replace: rename temp file to final file
            try:
                # Try to remove existing file (might fail if open in Excel)
                if output_file.exists():
                    output_file.unlink()
                
                temp_file.rename(output_file)
                
                logger.info(
                    f"EXCEL_EXPORT_SUCCESS: {output_file.name} "
                    f"(total_morning={data['total_morning']}, realtime={data['realtime']}, "
                    f"events={len(data['events'])}, alerts={len(data['alerts'])})"
                )
                return True
                
            except PermissionError:
                logger.warning(
                    f"EXCEL_EXPORT_SKIPPED: Cannot overwrite {output_file.name} - "
                    f"file may be open in Excel. Temp file preserved: {temp_file.name}"
                )
                # Keep temp file for debugging
                return False
            except Exception as e:
                logger.error(f"EXCEL_EXPORT_ERROR: Failed to rename temp file: {e}")
                if temp_file.exists():
                    temp_file.unlink()
                return False
        
        finally:
            conn.close()
            
    except sqlite3.Error as e:
        logger.error(f"EXCEL_EXPORT_ERROR: Database error: {e}", exc_info=True)
        if temp_file.exists():
            temp_file.unlink()
        return False
    except Exception as e:
        logger.error(f"EXCEL_EXPORT_ERROR: {e}", exc_info=True)
        if temp_file.exists():
            temp_file.unlink()
        return False


def _format_excel(wb):
    """Apply formatting to Excel workbook."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for sheet_name in ['SUMMARY', 'MISSING_PERIODS', 'ALERTS', 'EVENTS']:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Enable filter
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

